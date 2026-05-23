"""
Importa gastos desde el Excel 2026 (Plantilla-para-controlar-gastos.xlsm)
a la tabla Control_gastos en Supabase.

Uso:
    py finanzas_personales/db/importar_excel_control_gastos.py
    py finanzas_personales/db/importar_excel_control_gastos.py --dry-run
    py finanzas_personales/db/importar_excel_control_gastos.py --limpiar
"""

import sys
import re
from datetime import date
from pathlib import Path

from dotenv import load_dotenv
import os
import openpyxl
from supabase import create_client

load_dotenv(Path(__file__).parent.parent.parent / ".env")

EXCEL_PATH = Path(r"H:\Mi unidad\Finanza personal\Presupuesto familiar\2026\Plantilla-para-controlar-gastos.xlsm")
MESES = ["01 Enero", "02 Febrero", "03 Marzo", "04 Abril", "05 Mayo"]

SUPABASE_URL  = os.getenv("SUPABASE_FINANZAS_URL")
SUPABASE_KEY  = os.getenv("SUPABASE_FINANZAS_SERVICE_ROLE_KEY")
USER_ID       = os.getenv("FINANZAS_USER_ID")
TABLA         = "Control_gastos"


def detectar_forma_pago(detalle: str) -> str:
    if re.search(r'amis?pass', detalle, re.IGNORECASE):
        return "Amipass"
    return "Débito"


def evaluar_importe(valor) -> float | None:
    if valor is None:
        return None
    if isinstance(valor, (int, float)):
        return float(valor)
    texto = str(valor).strip()
    if texto.startswith("="):
        # Fórmula simple de suma: =A+B+C
        try:
            numeros = re.findall(r'\d+(?:\.\d+)?', texto)
            return float(sum(int(n) for n in numeros))
        except Exception:
            print(f"  WARN No se pudo evaluar fórmula: {texto} — fila ignorada")
            return None
    try:
        return float(texto.replace(".", "").replace(",", "."))
    except ValueError:
        return None


def leer_excel() -> list[dict]:
    print(f"Leyendo {EXCEL_PATH.name}...")
    wb = openpyxl.load_workbook(EXCEL_PATH, data_only=True, keep_vba=True)
    registros = []

    for mes in MESES:
        if mes not in wb.sheetnames:
            print(f"  WARN Hoja '{mes}' no encontrada, saltando")
            continue

        ws = wb[mes]
        fila_num = 7
        mes_count = 0

        for row in ws.iter_rows(min_row=8, values_only=True):
            fila_num += 1
            grupo   = row[1]
            concepto = row[2]
            fecha   = row[3]
            detalle = row[4]
            importe_raw = row[5]

            # Saltar filas sin datos esenciales
            if not grupo or not concepto or not importe_raw:
                continue

            importe = evaluar_importe(importe_raw)
            if importe is None or importe <= 0:
                continue

            # Fecha
            if isinstance(fecha, date):
                fecha_str = fecha.strftime("%Y-%m-%d")
            elif hasattr(fecha, 'date'):
                fecha_str = fecha.date().isoformat()
            else:
                print(f"  WARN Fecha inválida fila {fila_num} de {mes}: {fecha} — fila ignorada")
                continue

            detalle_str = str(detalle).strip() if detalle else ""
            forma_pago = detectar_forma_pago(detalle_str)

            registros.append({
                "user_id":    USER_ID,
                "tipo_tx":    "Gasto",
                "grupo":      str(grupo).strip(),
                "concepto":   str(concepto).strip(),
                "fecha_date": fecha_str,
                "detalle":    detalle_str,
                "importe":    importe,
                "forma_pago": forma_pago,
                "fuente":     "excel_2026",
            })
            mes_count += 1

        print(f"  OK {mes}: {mes_count} registros")

    return registros


def main():
    dry_run = "--dry-run" in sys.argv
    limpiar = "--limpiar" in sys.argv or True  # siempre limpia salvo que se indique lo contrario

    if not SUPABASE_URL or not SUPABASE_KEY or not USER_ID:
        print("ERROR Faltan variables de entorno: SUPABASE_FINANZAS_URL, SUPABASE_FINANZAS_SERVICE_ROLE_KEY, FINANZAS_USER_ID")
        sys.exit(1)

    registros = leer_excel()
    total = len(registros)
    amipass = sum(1 for r in registros if r["forma_pago"] == "Amipass")
    debito  = total - amipass

    print(f"\nTotal a insertar: {total} registros")
    print(f"  Débito:  {debito}")
    print(f"  Amipass: {amipass}")

    if dry_run:
        print("\n[DRY RUN] No se realizaron cambios en Supabase.")
        for r in registros[:5]:
            print(" ", r)
        print("  ...")
        return

    sb = create_client(SUPABASE_URL, SUPABASE_KEY)

    # Limpiar tabla
    print(f"\nLimpiando tabla {TABLA}...")
    sb.table(TABLA).delete().neq("id", 0).execute()
    print("  OK Tabla limpia")

    # Insertar en lotes de 100
    LOTE = 100
    insertados = 0
    for i in range(0, total, LOTE):
        lote = registros[i:i + LOTE]
        sb.table(TABLA).insert(lote).execute()
        insertados += len(lote)
        print(f"  Insertados {insertados}/{total}...")

    print(f"\nLISTO Migración completa: {insertados} registros en {TABLA}")


if __name__ == "__main__":
    main()
